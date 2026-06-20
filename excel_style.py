"""
excel_style.py — write a pandas DataFrame to a brand-styled .xlsx.

One function, used by every generator, so all spreadsheet outputs share the
Sea Green / Tahoma look defined in instructions.md (header fill, row banding,
hairline borders, right-aligned numeric columns, frozen header, autofit).
"""

from __future__ import annotations

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from common import BRAND, BRAND_FONT

_HAIRLINE = Side(style="thin", color=BRAND["hairline"])
_BORDER = Border(left=_HAIRLINE, right=_HAIRLINE, top=_HAIRLINE, bottom=_HAIRLINE)


def write_branded_excel(
    df: pd.DataFrame,
    path,
    sheet_name: str = "Data",
    numeric_cols: list[str] | None = None,
    title: str | None = None,
) -> None:
    """Write df to `path` as a styled workbook."""
    numeric_cols = numeric_cols or []
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        start_row = 1 if title else 0
        df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=start_row)
        ws = writer.sheets[sheet_name]

        ncols = len(df.columns)
        header_row = start_row + 1  # openpyxl is 1-indexed

        if title:
            ws.cell(row=1, column=1, value=title)
            tcell = ws.cell(row=1, column=1)
            tcell.font = Font(name=BRAND_FONT, bold=True, size=14,
                              color=BRAND["sea_green_deep"])
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)

        # Header styling
        header_fill = PatternFill("solid", fgColor=BRAND["sea_green_wash"])
        for c in range(1, ncols + 1):
            cell = ws.cell(row=header_row, column=c)
            cell.font = Font(name=BRAND_FONT, bold=True, color=BRAND["ink"])
            cell.fill = header_fill
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")

        # Body styling + row banding
        band = PatternFill("solid", fgColor=BRAND["row_band"])
        numeric_idx = {df.columns.get_loc(c) + 1 for c in numeric_cols
                       if c in df.columns}
        for r in range(header_row + 1, header_row + 1 + len(df)):
            banded = (r - header_row) % 2 == 0
            for c in range(1, ncols + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = Font(name=BRAND_FONT, color=BRAND["ink"])
                cell.border = _BORDER
                if banded:
                    cell.fill = band
                cell.alignment = Alignment(
                    horizontal="right" if c in numeric_idx else "left",
                    vertical="center",
                )

        # Number format on numeric columns
        for c in numeric_idx:
            for r in range(header_row + 1, header_row + 1 + len(df)):
                ws.cell(row=r, column=c).number_format = "#,##0"

        # Freeze header, autofit columns
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
        for c in range(1, ncols + 1):
            col = df.columns[c - 1]
            width = max(
                len(str(col)),
                int(df[col].astype(str).str.len().max() or 0),
            ) + 2
            ws.column_dimensions[get_column_letter(c)].width = min(width, 40)
